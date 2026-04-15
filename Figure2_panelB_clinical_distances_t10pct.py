import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# -----------------------------
# Inputs (Table 1 / SI Table S2)
# -----------------------------
drug_order = ["Edaravone", "Fasudil", "MgSO4", "Uric Acid", "NXY-059", "Minocycline"]
params = {
    "Edaravone": (342.39, 3.50e-05),
    "Fasudil": (288.44, 4.81e-04),
    "MgSO4": (387.28, 3.703e-05),
    "Uric Acid": (346.47, 1.44e-04),
    "NXY-059": (263.70, 3.83e-09),
    "Minocycline": (248.16, 8.37e-06),
}

L_cm_list = [1, 2, 5]     # slab thickness L (cm), so center is at L/2
threshold = 0.10         # 10%

# -----------------------------
# Analytic steady-state midpoint fraction (two-sided slab)
# Css_center/Cb = 1/cosh((L/2)/sqrt(D/k))
# -----------------------------
def css_center_slab(D_um2_s, k_s, L_um):
    if k_s <= 0:
        return 1.0
    lam = math.sqrt(D_um2_s / k_s)  # penetration length (µm)
    z = (L_um / 2.0) / lam
    if z > 700:
        return 0.0
    return 1.0 / math.cosh(z)

# -----------------------------
# Time-to-threshold under step boundary (explicit FD)
# Ct = D Cxx - k C, with C(0,t)=C(L,t)=1 and C(x,0)=0
# Uses N=501 nodes (same convention as SI code: dx = L/500)
# -----------------------------
def simulate_t_to_threshold(D_um2_s, k_s, L_um, target, N=501, t_end_h=200):
    Nx = N
    dx = L_um / (Nx - 1)
    dt = 0.9 / (k_s + 2.0 * D_um2_s / (dx * dx))  # explicit stability
    t_end_s = t_end_h * 3600.0
    n_steps = int(math.ceil(t_end_s / dt))

    C = np.zeros(Nx, dtype=float)
    C[0] = 1.0
    C[-1] = 1.0
    center = Nx // 2

    alpha = D_um2_s * dt / (dx * dx)
    t = 0.0
    for step in range(n_steps + 1):
        if C[center] >= target:
            return t / 3600.0  # hours
        if step == n_steps:
            break
        C_new = C.copy()
        C_new[1:-1] = (
            C[1:-1]
            + alpha * (C[2:] - 2.0 * C[1:-1] + C[:-2])
            - (k_s * dt) * C[1:-1]
        )
        C_new[0] = 1.0
        C_new[-1] = 1.0
        C = C_new
        t += dt

    return None  # did not cross within t_end_h (shouldn't happen if Css >= target)

# -----------------------------
# Compute table (blank if Css < threshold)
# -----------------------------
rows = []
for drug in drug_order:
    D, k = params[drug]
    row = {"Drug": drug}
    for L_cm in L_cm_list:
        L_um = L_cm * 10_000.0  # 1 cm = 10 mm = 10,000 µm
        css = css_center_slab(D, k, L_um)
        if css < threshold:
            row[f"L={L_cm} cm"] = None
        else:
            row[f"L={L_cm} cm"] = simulate_t_to_threshold(D, k, L_um, target=threshold, N=501, t_end_h=200)
    rows.append(row)

df = pd.DataFrame(rows)

# -----------------------------
# PNG plot (points only; no bars; missing = unreachable)
# -----------------------------
x = np.arange(len(drug_order))
plt.figure(figsize=(12, 4.8))
markers = {1: "o", 2: "s", 5: "D"}

for L_cm in L_cm_list:
    y = df[f"L={L_cm} cm"].values.astype(float)
    mask = ~np.isnan(y)
    plt.scatter(x[mask], y[mask], marker=markers[L_cm], label=f"L={L_cm} cm", s=80)

plt.xticks(x, drug_order, rotation=25, ha="right")
plt.ylabel(r"Time for centre to reach 10% of edge concentration, $t_{10\%}$ (hours)")
plt.title("Clinical-distance deliverability: time-to-10% at lesion centre (step boundary upper bound)")
plt.ylim(0, 100)
plt.legend(frameon=False, ncols=3, loc="upper left")
plt.tight_layout()
plt.savefig("/mnt/data/Fig2_clinical_distances_points_t10pct.png", dpi=300)
plt.close()

# -----------------------------
# Excel: simple data table + marker-only chart (editable)
# Strategy: reuse your existing 1% template file and overwrite values/titles
# -----------------------------
template = "/mnt/data/Fig2_clinical_distances_points_t1pct.xlsx"
wb = openpyxl.load_workbook(template)
ws = wb["data"]
ws_notes = wb["notes"]

# headers
ws["A1"].value = "Drug"
ws["B1"].value = "L=1 cm"
ws["C1"].value = "L=2 cm"
ws["D1"].value = "L=5 cm"

# data
for i, drug in enumerate(drug_order, start=2):
    ws.cell(row=i, column=1).value = drug
    for j, L_cm in enumerate(L_cm_list, start=2):
        val = df.loc[df["Drug"] == drug, f"L={L_cm} cm"].values[0]
        if isinstance(val, float) and math.isnan(val):
            ws.cell(row=i, column=j).value = None
        else:
            ws.cell(row=i, column=j).value = float(val) if val is not None else None

# notes
ws_notes["A3"].value = (
    "Each point is t10% = first time the lesion centre reaches 10% of the edge concentration "
    "under a step edge concentration (optimistic upper bound)."
)
ws_notes["A4"].value = (
    "Missing points mean the 10% centre threshold is analytically unreachable at that L "
    "(Css_center < 10%) under the diffusion–consumption model."
)

# update chart titles (chart already exists in template)
chart = ws._charts[0]
chart.title = "Time-to-10% at lesion centre (step boundary upper bound)"
chart.y_axis.title = "t10% (hours)"
chart.x_axis.title = "Drug"

wb.save("/mnt/data/Fig2_clinical_distances_points_t10pct.xlsx")
