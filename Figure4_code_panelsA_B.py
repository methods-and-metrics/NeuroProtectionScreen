#!/usr/bin/env python3
"""
Code to generate main Figure 4 (Panels A and B).

Panel A: effect of time-varying boundary exposure on centre delivery
- Drug: minocycline (Table 1 / SI Table S2)
- Geometry: two-sided 1D slab, L = 5 cm (centre-to-rim distance d = 25 mm)
- Model: Ct = D Cxx - k C
- Boundary: (i) step (constant edge concentration = 1; optimistic upper bound)
           (ii) exponential decay with half-life 18 h (illustrative finite systemic exposure)

Panel B: clinical exemplar bar chart (TASTE-2)
- Outcome: functional independence (mRS 0–2) at 90 days, % (overall ITT and mismatch subgroup)

Outputs:
- Panel A PNG: /mnt/data/Fig4_panelA_time_varying_boundary_minocycline.png
- Panel B PNG: /mnt/data/Fig4_panelB_clinical_exemplar_TASTE2.png
- Panel B Excel: /mnt/data/Fig4_panelB_clinical_exemplar_TASTE2.xlsx
"""

import math
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# -----------------------------
# Panel A: minocycline PK-decay example
# -----------------------------

def simulate_center_two_sided_slab(
    D_um2_s: float,
    k_s: float,
    L_um: float,
    dx_um: float = 100.0,
    t_end_h: float = 200.0,
    boundary_func=None,
    sample_every_s: float = 600.0,  # 10 min
):
    """Explicit FD for Ct = D*Cxx - k*C with two-sided Dirichlet boundary C(0,t)=C(L,t)=Cb(t).
    Returns (time_h, C_center(t), C_boundary(t)), all normalised to initial boundary (Cb(0)=1 assumed)."""
    if boundary_func is None:
        boundary_func = lambda t: 1.0

    nx = int(round(L_um / dx_um)) + 1
    dx = L_um / (nx - 1)

    # stability-controlled dt
    dt = 0.9 / (k_s + 2.0 * D_um2_s / (dx * dx))
    t_end_s = t_end_h * 3600.0
    n_steps = int(math.ceil(t_end_s / dt))
    sample_every = max(1, int(round(sample_every_s / dt)))

    alpha = D_um2_s * dt / (dx * dx)
    if alpha > 0.5:
        raise ValueError(f"Unstable scheme: alpha={alpha:.3f} > 0.5")

    C = np.zeros(nx, dtype=float)
    C[0] = boundary_func(0.0)
    C[-1] = boundary_func(0.0)
    mid = nx // 2

    t_out, c_mid, b_out = [], [], []
    for step in range(n_steps + 1):
        t = step * dt
        b = boundary_func(t)

        if step % sample_every == 0 or step == n_steps:
            t_out.append(t / 3600.0)
            c_mid.append(C[mid])
            b_out.append(b)

        if step == n_steps:
            break

        C_new = C.copy()
        C_new[1:-1] = (
            C[1:-1]
            + alpha * (C[2:] - 2.0 * C[1:-1] + C[:-2])
            - (k_s * dt) * C[1:-1]
        )
        C_new[0] = b
        C_new[-1] = b
        C = C_new

    return np.array(t_out), np.array(c_mid), np.array(b_out)

def panelA_make_plot():
    # Minocycline parameters (Table 1 / SI Table S2)
    D = 248.16   # µm^2/s
    k = 8.37e-6  # 1/s
    L_um = 50_000.0  # 5 cm

    # Step boundary (upper bound)
    t_h, c_step, _ = simulate_center_two_sided_slab(D, k, L_um, boundary_func=lambda t: 1.0)

    # Exponential boundary decay (illustrative finite systemic exposure)
    t_half_h = 18.0
    tau = (t_half_h * 3600.0) / math.log(2.0)
    boundary_decay = lambda t: math.exp(-t / tau)

    t_h, c_decay, b_decay = simulate_center_two_sided_slab(D, k, L_um, boundary_func=boundary_decay)

    # Plot
    plt.figure(figsize=(7.2, 4.8))
    plt.plot(t_h, c_step, label="Centre (step boundary; upper bound)")
    plt.plot(t_h, c_decay, label=f"Centre (boundary decays, t1/2={t_half_h:.0f} h)")
    plt.plot(t_h, b_decay, linestyle="--", label=f"Boundary (decays, t1/2={t_half_h:.0f} h)")
    plt.axhline(0.01, linestyle=":", label="1% threshold")

    plt.yscale("log")
    plt.xlabel("Time (h)")
    plt.ylabel("Normalised concentration (C / C_boundary(0))")
    plt.title("Effect of time-varying boundary exposure on centre delivery\nMinocycline, L=5 cm (d=25 mm), diffusion–reaction model")
    plt.xlim(0, 200)
    plt.ylim(1e-4, 2)

    plt.legend(frameon=False, fontsize=8, loc="upper right")
    plt.tight_layout()
    plt.savefig("/mnt/data/Fig4_panelA_time_varying_boundary_minocycline.png", dpi=300)
    plt.close()

# -----------------------------
# Panel B: TASTE-2 clinical exemplar bar chart (+ Excel)
# -----------------------------

def panelB_make_png_and_excel():
    # Values reproduced from the trial report / SI extraction
    groups = ["Overall (ITT)", "Mismatch subgroup"]
    treatment = [55.0, 55.5]  # Edaravone–dexborneol
    placebo = [49.6, 42.9]

    annotation = "Overall: RR 1.11; RD 5.4%; p=0.05\nMismatch: RR 1.29; RD 13.0%; Pinteraction=0.003"

    # --- PNG ---
    x = np.arange(len(groups))
    width = 0.36

    plt.figure(figsize=(7.6, 4.8))
    bars1 = plt.bar(x - width/2, treatment, width, label="Edaravone–dexborneol")
    bars2 = plt.bar(x + width/2, placebo, width, label="Placebo")

    plt.ylabel("Functional independence (mRS 0–2 at 90 days), %")
    plt.ylim(0, 80)
    plt.xticks(x, groups)
    plt.title("Clinical exemplar: stronger signal in imaging-favourable stratum (TASTE-2)")

    # value labels
    for b in list(bars1) + list(bars2):
        h = b.get_height()
        plt.text(b.get_x() + b.get_width()/2, h + 1.0, f"{h:.1f}%", ha="center", va="bottom", fontsize=9)

    # annotation to the right
    plt.text(1.55, 40, annotation, ha="left", va="center", fontsize=9)

    plt.legend(frameon=False, loc="upper left")
    plt.tight_layout()
    plt.savefig("/mnt/data/Fig4_panelB_clinical_exemplar_TASTE2.png", dpi=300)
    plt.close()

    # --- Excel ---
    xlsx_path = "/mnt/data/Fig4_panelB_clinical_exemplar_TASTE2.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Group", "Edaravone–dexborneol", "Placebo"])
    for g, t, p in zip(groups, treatment, placebo):
        ws.append([g, t, p])

    # annotation cells
    ws["E2"].value = "Annotation"
    ws["E3"].value = annotation

    # chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = "Clinical exemplar: stronger signal in imaging-favourable stratum (TASTE-2)"
    chart.y_axis.title = "mRS 0–2 at 90 days (%)"
    chart.x_axis.title = ""
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 80

    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1 + len(groups))
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(groups))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "A6")
    wb.save(xlsx_path)

if __name__ == "__main__":
    panelA_make_plot()
    panelB_make_png_and_excel()
    print("Wrote Panel A PNG and Panel B PNG/Excel to /mnt/data")
